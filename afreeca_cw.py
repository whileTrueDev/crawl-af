
# coding: utf-8

# In[59]:


import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import warnings
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
write_wb = Workbook()
global rank #rank를 전역변수로 설정
global i #i를 전역변수로 설정
i = 0
rank = 2

warnings.filterwarnings("ignore", category=DeprecationWarning) #경고 메세지 무시

stime = time.time() #현재 시각을 stime으로 받기

# Headless
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
options.add_argument("lang=ko_KR")

now = datetime.now() #크롤링 시작시간을 간단하게 보기위한 함수 now
print("크롤링 시작시간은",now)

write_ws = write_wb.active #엑셀파일에 쓰기

write_ws['A1'] = 'BJ 이름'
write_ws['B1'] = 'BJ_id'
write_ws['C1'] = '방송제목'
write_ws['D1'] = '현재 시청자수'
write_ws['E1'] = '방송 시작시간'
write_ws['F1'] = '카테고리'
write_ws['G1'] = '데이터가 크롤링된 시간'
title_i = [] #방송제목
name_i = [] #bj이름
viewer_i = [] #현재 시청자수
start_time_i = [] #방송 시작시간
category_i = [] #방송 카테고리
current_time_i = [] #현재 시각
bj_id_i = [] #bj의id
write_wb.save('방송데이터.xlsx') #방송 데이터라는 이름의 엑셀 파일로 저장
rb = load_workbook(r"C:\Users\user\works\방송데이터.xlsx")

driver = webdriver.Chrome('C:/chromedriver.exe', options=options) #크롬 드라이버 실행하기
driver.get('http://www.afreecatv.com')


# In[60]:


def open_driver(): # 아프리카 홈페이지에서 100명까지 방송 들어가기위한 셀레니움 이용한 클릭
    driver.find_element_by_css_selector("""#broadlist_area > div > ul > li:nth-child({}) > div > a.box_link""".format(rank-1)).click()
    driver.switch_to_window(driver.window_handles[1])
    driver.get(driver.current_url)


# In[61]:


def save_bj_name(): # 방송 안에서 bj의 닉네임을 가져오기
    bj_name = driver.find_element_by_css_selector("""#player_area > div.broadcast_information > div.text_information > div.nickname""")
    name_i.append(bj_name.text)
    write_ws['A{}'.format(rank)] = bj_name.text #A행에 bj닉네임 등록


# In[62]:


def save_title(): #live화면에서 방송제목 가져오기
    title = driver.find_element_by_css_selector("""#player_area > div.broadcast_information > div.text_information > div.broadcast_title > span""")
    title_i.append(title.text)
    write_ws['C{}'.format(rank)] = title.text #B행에 방송제목 등록


# In[63]:


def current_viewer(): #live화면에서 현재 시청자수 가져오기
    driver.find_element_by_xpath("""//*[@id="stop_screen"]/dl/dd[2]/a""").click()
    time.sleep(4)
    viewer = driver.find_element_by_xpath("""//*[@id="nAllViewer"]""")
    viewer_i.append(viewer.text)
    if str(0) in viewer.text: #크롤링중 방송이 종료되었을때 데이터 수집 안함
        write_ws['D{}'.format(rank)] = '방송이 종료되었습니다' 
    else:
        write_ws['D{}'.format(rank)] = viewer.text #방송이 종료되지 않을 경우 현재 시청자수 수집(D행)


# In[64]:


def save_start_time():#방송 시작시간 live화면에서 가져오기
    start_time = driver.find_element_by_css_selector("""#player_area > div.broadcast_information > div.text_information > ul > li:nth-child(1) > span""")
    start_time_i.append(start_time.text)
    write_ws['E{}'.format(rank)] = start_time.text #E행에 방송 시작시간 등록


# In[65]:


def save_category(): #live화면에서 카테고리 가져오기
    category =  driver.find_element_by_css_selector("""#player_area > div.broadcast_information > div.text_information > ul > li:nth-child(4) > span""")
    category_i.append(category.text)
    write_ws['F{}'.format(rank)] = category.text  #F행에 카테고리 등록


# In[66]:


def screenshot(): #각 라이브 화면을 스크린샷으로 찍음
    driver.save_screenshot("bj_{}.png".format(rank-1))


# In[67]:


def bj_id(): #bj id B행에 등록
    info_id = driver.current_url.split('/') #bj의 id를 방송국에 가서 가져오면 시간이 더걸릴거같아 url에서 id 추출
    bj_id_i.append(info_id[3])
    write_ws['B{}'.format(rank)] = info_id[3] #url에서 /으로 나눈 3번째에 id가 있어서 그id B행에 등록


# In[68]:


def current_time(): # 수집한 현재시간 
    now_i = datetime.now()
    current_time_i.append(now_i)
    write_ws['G{}'.format(rank)] = now_i # 수집한 현재시간 G행에 등록


# In[69]:


while rank<102: #100번 실행하도록 하는 반복문
    open_driver()
    save_bj_name()
    bj_id()
    save_title()
    current_viewer()
    save_start_time()
    save_category()
    screenshot()
    driver.close()
    current_time()
    driver.switch_to_window(driver.window_handles[0])
    rank += 1
    if not rank % 61: #live 방송수가 60번째가되면 더보기를 클릭해야 100번까지 볼수있으므로 더보기 클릭
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        driver.find_element_by_css_selector("""#broadlist_area > div > div > a""").click()
        time.sleep(3)
    if rank == 102: #102번째가되면 드라이버 종료
        driver.quit()

write_wb.save('방송데이터.xlsx') #지금까지 데이터 방송 데이터에 수집
now = datetime.now()
print("크롤링 끝난시간", now ) #크롤링이 끝난 시간을 반환
print("\n진행시간 {}초".format(time.time() - stime))


# In[70]:


bj_list = ['khm11903','superbsw123','qweqwe','123qd']  #인자로 넣을 리스트의 예시


# In[71]:


a = len(bj_list) #반복문에 리스트에 id 수만큼 반복할수 있도록 id수 세기


# In[76]:


while i < a: 
    data = pd.read_excel('방송데이터.xlsx') #방송 데이터에서 데이터 읽기
    int(i)
    if bj_list[i] in bj_id_i: #bj_list속 id가 bj_id_i에 있는지 비교하기
        indexno = bj_id_i.index(bj_list[i]) #있다면 인덱스 번호 반환
        print(data.loc[indexno])#방송데이터 엑셀에서 인덱스 번호 찾고 출력
    else:
        print("해당 bj가 방송중이지 않습니다")
    i = i + 1

