
# coding: utf-8

# In[1]:


import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import warnings
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
write_wb = Workbook()
global rank

rank = 2

warnings.filterwarnings("ignore", category=DeprecationWarning) 

stime = time.time()

write_ws = write_wb.create_sheet('test')

# Headless
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
options.add_argument("lang=ko_KR")

now = datetime.now()
print("크롤링 시작시간은",now)

write_ws = write_wb.active

write_ws['A1'] = 'BJ 이름'
write_ws['B1'] = 'BJ id'
write_ws['C1'] = '방송제목'
write_ws['D1'] = '현재 시청자수'
write_ws['E1'] = '방송 시작시간'
write_ws['F1'] = '카테고리'
write_ws['G1'] = '데이터가 크롤링된 시간'
title_i = []
name_i = []
viewer_i = []
start_time_i = []
category_i = []
current_time_i = []
bj_id_i = []
prin_i = []
write_wb.save('방송데이터.xlsx')
rb = load_workbook(r"C:\Users\user\works\방송데이터.xlsx")

driver = webdriver.Chrome('C:/chromedriver.exe', options=options) #크롬 드라이버 실행하기
driver.get('http://www.afreecatv.com')


# In[2]:


def open_driver():
    driver.find_element_by_xpath("""//*[@id="broadlist_area"]/div/ul/li[{}]/div/a[1]""".format(rank-1)).click()
    driver.switch_to_window(driver.window_handles[1])
    driver.get(driver.current_url)


# In[3]:


def save_bj_name():
    bj_name = driver.find_element_by_xpath("""//*[@id="player_area"]/div[2]/div[2]/div[1]""")
    name_i.append(bj_name.text)
    write_ws['A{}'.format(rank)] = bj_name.text    


# In[4]:


def save_title():
    title = driver.find_element_by_css_selector("""#player_area > div.broadcast_information > div.text_information > div.broadcast_title > span""")
    title_i.append(title.text)
    write_ws['C{}'.format(rank)] = title.text


# In[5]:


def current_viewer():
    driver.find_element_by_xpath("""//*[@id="stop_screen"]/dl/dd[2]/a""").click()
    time.sleep(4)
    viewer = driver.find_element_by_xpath("""//*[@id="nAllViewer"]""")
    viewer_i.append(viewer.text)
    if str(0) in viewer.text:
        write_ws['D{}'.format(rank)] = '방송이 종료되었습니다'
    write_ws['D{}'.format(rank)] = viewer.text


# In[6]:


def save_start_time():
    start_time = driver.find_element_by_css_selector("""#player_area > div.broadcast_information > div.text_information > ul > li:nth-child(1) > span""")
    start_time_i.append(start_time.text)
    write_ws['E{}'.format(rank)] = start_time.text


# In[7]:


def screenshot():
    driver.save_screenshot("bj_{}.png".format(rank-1))


# In[8]:


def save_category():
    category =  driver.find_element_by_css_selector("""#player_area > div.broadcast_information > div.text_information > ul > li:nth-child(4) > span""")
    category_i.append(category.text)
    write_ws['F{}'.format(rank)] = category.text


# In[9]:


def bj_id():
    info_id = driver.current_url.split('/')
    bj_id_i.append(info_id[3])
    write_ws['B{}'.format(rank)] = info_id[3]


# In[10]:


def current_time():
    now_i = datetime.now()
    current_time_i.append(now_i)
    write_ws['G{}'.format(rank)] = now_i


# In[11]:


def prin():
    aa = driver.find_element_by_xpath("""//*[@id="broadlist_area"]/div/ul/li[{}]/div/a[2]""".format(rank-1))
    prin_i.append(aa.text)


# In[12]:


while rank<102:
    open_driver()
    save_bj_name()
    bj_id()
    save_title()
    current_viewer()
    save_start_time()
    save_category()
    screenshot()
    driver.close()
    current_time()
    driver.switch_to_window(driver.window_handles[0])
    prin()
    rank += 1
    if not rank % 61:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        driver.find_element_by_css_selector("""#broadlist_area > div > div > a""").click()
        time.sleep(3)
    if rank == 102:
        driver.quit()
write_wb.save('방송데이터.xlsx')
now = datetime.now()
print("크롤링 끝난시간", now )
print("\n진행시간 {}초".format(time.time() - stime))


# In[24]:


def search_i():
    data = pd.read_excel('방송데이터.xlsx')
    if search in prin_i:
        indexno = prin_i.index(search)
        print(data.loc[indexno])
    else :
        if search == 'quit':
            print("종료")
        else:
            print("해당 bj가 존재하지 않습니다")


# In[25]:


while True:
    search  = input("검색을 원하는 닉네임을 입력해주세요(quit 입력시 검색종료)")
    search_i()
    if search == 'quit':
        break;

